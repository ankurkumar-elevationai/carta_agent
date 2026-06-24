"""
Carta Extraction Pipeline - Comprehensive Schema & Entity Analysis
Scans all JSON files from all export runs and produces an Excel workbook with:
  1. Entity Inventory (Funds, SPVs, Portfolio Companies, Investors, Securities, Valuations)
  2. Complete Field Catalog (field name, type, source, example, description)
  3. Platform Mapping Matrix (source field → target object, relationship type)
  4. ERD Relationships (entity relationships derived from the data)
"""

import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────
EXPORTS_ROOT = Path(r"c:\Users\iaman\Vscode Pycharm\openclaw_carta\output\exports")
OUTPUT_FILE = Path(r"c:\Users\iaman\Vscode Pycharm\openclaw_carta\output\carta_extraction_analysis.xlsx")

# ── Styles ──────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── Entity Classification ───────────────────────────────────────────
ENTITY_TYPE_MAP = {
    "fund": "Fund",
    "spv": "SPV",
    "investment": "Portfolio Company",
    "partner": "Investor/LP",
    "gp_entity": "GP Entity",
    "fund_family": "Fund Family",
    "management_company": "Management Company",
}

SOURCE_TYPE_MAP = {
    "Company Investment": "Portfolio Company",
    "SPV": "SPV",
    "Syndicate SPV": "SPV",
    "Fund": "Fund",
    "Partner": "Investor/LP",
    "GP Entity": "GP Entity",
    "Fund Family": "Fund Family",
    "Management Company": "Management Company",
}

# Entity types from entity_list that are actually config flags, not real entities
NOISE_ENTITY_TYPES = {"default", "firm"}

# ── Field Description Heuristics ────────────────────────────────────
FIELD_DESCRIPTIONS = {
    "id": "Unique identifier",
    "pk": "Primary key",
    "pk_key": "Primary key field name",
    "name": "Display name",
    "fund-name": "Fund display name",
    "fund_name": "Fund display name",
    "fund-id": "Fund identifier",
    "firm-id": "Firm identifier",
    "firm-uuid": "Firm UUID",
    "entity_id": "Entity identifier",
    "entity_name": "Entity display name",
    "entity_type": "Entity classification type",
    "org_pk": "Organization primary key",
    "source": "Entity source/origin type",
    "source_url": "API endpoint URL",
    "category": "Data category classification",
    "status": "Current status",
    "status_code": "HTTP response status code",
    "label": "Certificate/security label",
    "issue_date": "Date of issuance",
    "issuable_type": "Type of security (Common, Preferred, etc.)",
    "stock_type": "Stock classification",
    "quantity": "Number of shares/units",
    "cost": "Cost basis",
    "value": "Current value",
    "currency": "Currency code",
    "cost_to_exercise": "Cost to exercise options",
    "exercise_type": "Option exercise type",
    "has_vesting": "Whether vesting schedule exists",
    "is_vesting": "Whether currently vesting",
    "is_realized": "Whether investment is realized",
    "is_canceled": "Whether canceled",
    "is_pending": "Whether pending",
    "is_expired": "Whether expired",
    "qsbs": "QSBS qualification status",
    "landing_url": "Navigation URL for entity",
    "purpose": "Entity purpose/description",
    "children": "Child entities",
    "entity_group_id": "Entity group identifier",
    "entity_group_name": "Entity group name",
    "latency_ms": "API response latency in milliseconds",
    "shape_hash": "Response schema fingerprint",
    "replay_strategy": "API replay method used",
    "capability_tags": "Extracted capability tags",
    "x_carta_trace_id": "Carta internal trace ID",
    "capture_timestamp": "Time of data capture",
    "traversal_session_id": "Browser traversal session ID",
    "owner_id": "Owner entity identifier",
    "accepted_by_holder": "Whether accepted by the holder",
    "sent_to_holder": "Whether sent to holder",
    "sent_to_officer": "Whether sent to officer",
    "is_signing_state": "Whether in signing state",
    "committed": "Capital committed",
    "contributed": "Capital contributed",
    "distributed": "Capital distributed",
    "exercised": "Options exercised count",
    "exercisable": "Options exercisable count",
    "vested": "Shares/options vested",
    "settled": "Shares settled",
    "remaining_shares": "Remaining shares",
    "eligible_for_settlement": "Shares eligible for settlement",
    "time_vested": "Time-based vesting count",
    "threshold_value": "Threshold value for carried interest",
    "quantity_canceled": "Quantity of canceled shares",
    "investments-tab-enabled": "Whether investments tab is enabled",
    "general-ledger-enabled": "Whether general ledger is enabled",
    "in-app-valuations-enabled": "Whether in-app valuations enabled",
    "is_staff": "Whether user is staff",
    "show_edit_value": "Whether edit value is shown",
    "show_request_delivery": "Whether delivery request shown",
    "requires_two_person": "Whether two-person approval required",
    "original_acquisition_date": "Original acquisition date",
    "is_certificated": "Whether share is certificated",
    "cost_includes_debt_canceled": "Whether cost includes canceled debt",
    "is_any_paper": "Whether any paper certificates exist",
    "is_fully_exercised": "Whether fully exercised",
    "is_terminated": "Whether terminated",
    "is_withdrawn": "Whether withdrawn",
    "products": "Available platform products",
}

# ── Platform Mapping Targets ────────────────────────────────────────
PLATFORM_OBJECTS = {
    "fund-name": ("Fund", "attribute", "Required"),
    "fund-id": ("Fund", "identifier", "Required"),
    "firm-id": ("Firm", "identifier", "Required"),
    "firm-uuid": ("Firm", "identifier", "Required"),
    "fund-admin-fund-pk": ("Fund Admin", "foreign_key", "Optional"),
    "entity_id": ("Entity", "identifier", "Required"),
    "entity_name": ("Entity", "attribute", "Required"),
    "entity_type": ("Entity", "classifier", "Required"),
    "org_pk": ("Organization", "foreign_key", "Required"),
    "id": ("Record", "identifier", "Required"),
    "pk_key": ("Record", "identifier_type", "Required"),
    "label": ("Security", "attribute", "Required"),
    "issue_date": ("Security", "attribute", "Required"),
    "issuable_type": ("Security", "classifier", "Required"),
    "status": ("Security", "attribute", "Required"),
    "quantity": ("Holding", "metric", "Required"),
    "cost": ("Holding", "metric", "Optional"),
    "value": ("Holding", "metric", "Optional"),
    "currency": ("Holding", "attribute", "Required"),
    "stock_type": ("Security", "classifier", "Optional"),
    "owner_id": ("Investor", "foreign_key", "Required"),
    "name": ("Entity", "attribute", "Required"),
    "source": ("Entity", "classifier", "Required"),
    "landing_url": ("Navigation", "reference", "Optional"),
    "purpose": ("Entity", "attribute", "Optional"),
    "children": ("Entity", "relationship", "Optional"),
    "has_vesting": ("Security", "attribute", "Optional"),
    "is_realized": ("Holding", "attribute", "Optional"),
    "qsbs": ("Tax", "attribute", "Optional"),
    "committed": ("Fund Accounting", "metric", "Optional"),
    "contributed": ("Fund Accounting", "metric", "Optional"),
    "distributed": ("Fund Accounting", "metric", "Optional"),
    "exercised": ("Security", "metric", "Optional"),
    "exercisable": ("Security", "metric", "Optional"),
    "vested": ("Security", "metric", "Optional"),
    "products": ("Platform", "attribute", "Optional"),
    "source_url": ("API", "reference", "Required"),
    "category": ("API", "classifier", "Required"),
    "capability_tags": ("API", "classifier", "Optional"),
    "replay_strategy": ("API", "attribute", "Optional"),
    "latency_ms": ("API", "metric", "Optional"),
    "shape_hash": ("Schema", "identifier", "Optional"),
}


def infer_type(val):
    """Infer a human-readable type string from a Python value."""
    if val is None:
        return "null"
    if isinstance(val, bool):
        return "boolean"
    if isinstance(val, int):
        return "integer"
    if isinstance(val, float):
        return "float"
    if isinstance(val, str):
        if re.match(r"^\d{2}/\d{2}/\d{4}$", val) or re.match(r"^\d{4}-\d{2}-\d{2}", val):
            return "date"
        if re.match(r"^[0-9a-f]{8}-", val):
            return "uuid"
        if val.startswith("/") or val.startswith("http"):
            return "url"
        return "string"
    if isinstance(val, list):
        return "array"
    if isinstance(val, dict):
        return "object"
    return type(val).__name__


def truncate(val, max_len=80):
    """Truncate a value for display."""
    s = str(val)
    return s[:max_len] + "..." if len(s) > max_len else s


def flatten_keys(obj, prefix="", depth=0, max_depth=3):
    """Recursively flatten JSON keys up to max_depth."""
    results = {}
    if depth > max_depth:
        return results
    if isinstance(obj, dict):
        for k, v in obj.items():
            full_key = f"{prefix}.{k}" if prefix else k
            results[full_key] = v
            if isinstance(v, (dict, list)) and depth < max_depth:
                results.update(flatten_keys(v, full_key, depth + 1, max_depth))
    elif isinstance(obj, list) and obj:
        # Sample first element
        results.update(flatten_keys(obj[0], f"{prefix}[0]", depth + 1, max_depth))
    return results


def style_header(ws, row_num, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=12, max_width=50):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ══════════════════════════════════════════════════════════════════════
# MAIN ANALYSIS
# ══════════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("Carta Extraction Pipeline - Schema & Entity Analysis")
    print("=" * 60)

    # ── Phase 1: Scan all JSON files ────────────────────────────────
    entities = {}          # entity_id → {name, type, category, source, ...}
    field_catalog = {}     # field_path → {type, category, example, count}
    all_files = []
    manifests = []

    for run_dir in EXPORTS_ROOT.iterdir():
        if not run_dir.is_dir():
            continue

        # Load manifests
        manifest_path = run_dir / "extracted" / "_extraction_manifest.json"
        if manifest_path.exists():
            with open(manifest_path, "r", encoding="utf-8") as f:
                manifests.append(json.load(f))

        # Scan extracted JSONs
        extracted_dir = run_dir / "extracted"
        if not extracted_dir.exists():
            continue

        for json_file in extracted_dir.rglob("*.json"):
            if json_file.name == "_extraction_manifest.json":
                continue
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    doc = json.load(f)
            except (json.JSONDecodeError, UnicodeDecodeError):
                continue

            meta = doc.get("_meta", {})
            data = doc.get("data", doc)
            category = meta.get("category", json_file.parent.parent.name)
            entity_id = meta.get("entity_id", "")
            entity_name = meta.get("entity_name", "")
            entity_type = meta.get("entity_type", "")

            all_files.append({
                "file": str(json_file.relative_to(EXPORTS_ROOT)),
                "category": category,
                "entity_id": entity_id,
                "entity_name": entity_name,
                "entity_type": entity_type,
                "source_url": meta.get("source_url", ""),
                "status_code": meta.get("status_code", ""),
            })

            # Register entity
            if entity_id and entity_id not in entities:
                entities[entity_id] = {
                    "entity_id": entity_id,
                    "entity_name": entity_name,
                    "entity_type": entity_type,
                    "classified_as": ENTITY_TYPE_MAP.get(entity_type, entity_type),
                    "category": category,
                    "org_pk": meta.get("org_pk", ""),
                    "file_count": 0,
                    "source_urls": set(),
                }
            if entity_id in entities:
                entities[entity_id]["file_count"] += 1
                if meta.get("source_url"):
                    entities[entity_id]["source_urls"].add(meta.get("source_url"))

            # Catalog fields from data
            if isinstance(data, dict):
                flat = flatten_keys(data, prefix="data", max_depth=2)
            elif isinstance(data, list) and data:
                flat = flatten_keys(data[0], prefix="data[0]", max_depth=2)
            else:
                flat = {}

            for field_path, val in flat.items():
                if field_path not in field_catalog:
                    field_catalog[field_path] = {
                        "type": infer_type(val),
                        "category": category,
                        "example": truncate(val),
                        "count": 0,
                        "categories": set(),
                    }
                field_catalog[field_path]["count"] += 1
                field_catalog[field_path]["categories"].add(category)

        # Also scan entity list files from valuations (gp-firm-entity-list)
        for json_file in (run_dir / "extracted").rglob("*.json"):
            if json_file.name == "_extraction_manifest.json":
                continue
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    doc = json.load(f)
            except:
                continue
            data = doc.get("data", [])
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict) and "name" in item and "source" in item:
                        raw_type = item.get("source", "unknown").lower().replace(" ", "_")
                        # Skip config flags that aren't real entities
                        if raw_type in NOISE_ENTITY_TYPES:
                            continue
                        eid = item.get("name", "").replace(" ", "_").lower()
                        if eid and eid not in entities:
                            entities[eid] = {
                                "entity_id": eid,
                                "entity_name": item["name"],
                                "entity_type": raw_type,
                                "classified_as": SOURCE_TYPE_MAP.get(item.get("source", ""), "Other"),
                                "category": "entity_list",
                                "org_pk": "",
                                "file_count": 1,
                                "source_urls": set(),
                            }

    # ── Deduplicate entities by name (keep the one with most files) ──
    seen_names = {}
    deduped_entities = {}
    for eid, e in entities.items():
        name = e["entity_name"]
        if name not in seen_names or e["file_count"] > seen_names[name]["file_count"]:
            seen_names[name] = e
            deduped_entities[eid] = e
        # Remove any leftover duplicate keys
    entities = {}
    for name, e in seen_names.items():
        entities[e["entity_id"]] = e

    # ── Filter noisy fields from field catalog ──
    def is_noisy_field(fp, info):
        leaf = fp.split(".")[-1].split("[")[0]
        # Pure numeric keys (fund IDs used as dict keys)
        if leaf.isdigit():
            return True
        # Internal API/URL path fields
        if info["type"] == "url" and any(x in fp for x in ["-url", "_url", "api-url", "api_url"]):
            return True
        # Platform internals
        if any(x in fp for x in ["is-generate-upload", "is-persist-document", "pendo", "csrf", "datadog"]):
            return True
        return False

    clean_catalog = {fp: info for fp, info in field_catalog.items() if not is_noisy_field(fp, info)}
    removed = len(field_catalog) - len(clean_catalog)
    field_catalog = clean_catalog

    print(f"  Scanned {len(all_files)} JSON files across {len(list(EXPORTS_ROOT.iterdir()))} runs")
    print(f"  Found {len(entities)} unique entities (deduped by name)")
    print(f"  Cataloged {len(field_catalog)} business fields ({removed} noise fields removed)")

    # ── Phase 2: Build Excel Workbook ───────────────────────────────
    wb = Workbook()

    # ────────────────────────────────────────────────────────────────
    # SHEET 1: Entity Inventory
    # ────────────────────────────────────────────────────────────────
    ws_inv = wb.active
    ws_inv.title = "Entity Inventory"
    headers = ["Entity ID", "Entity Name", "Raw Type", "Classification",
                "Source Category", "Org PK", "File Count", "API Endpoints"]
    ws_inv.append(headers)
    style_header(ws_inv, 1, len(headers))

    # Group by classification
    classified = defaultdict(list)
    for e in entities.values():
        classified[e["classified_as"]].append(e)

    row = 2
    for cls_name in ["Fund", "Fund Family", "SPV", "Portfolio Company", "Investor/LP", "GP Entity", "Management Company"]:
        items = classified.get(cls_name, [])
        if not items:
            continue
        # Section header
        ws_inv.cell(row=row, column=1, value=f"── {cls_name} ({len(items)}) ──")
        ws_inv.cell(row=row, column=1).font = Font(bold=True, size=11, color="2F5496")
        for c in range(1, len(headers) + 1):
            ws_inv.cell(row=row, column=c).fill = SUBHEADER_FILL
        row += 1

        for e in sorted(items, key=lambda x: x["entity_name"]):
            ws_inv.append([
                e["entity_id"],
                e["entity_name"],
                e["entity_type"],
                e["classified_as"],
                e["category"],
                e["org_pk"],
                e["file_count"],
                len(e["source_urls"]),
            ])
            row += 1

    auto_width(ws_inv)

    # ────────────────────────────────────────────────────────────────
    # SHEET 2: Field Catalog
    # ────────────────────────────────────────────────────────────────
    ws_fields = wb.create_sheet("Field Catalog")
    headers = ["Field Path", "Data Type", "Source Categories", "Occurrences",
                "Example Value", "Description"]
    ws_fields.append(headers)
    style_header(ws_fields, 1, len(headers))

    for fp in sorted(field_catalog.keys()):
        info = field_catalog[fp]
        leaf = fp.split(".")[-1].split("[")[0]
        desc = FIELD_DESCRIPTIONS.get(leaf, "")
        ws_fields.append([
            fp,
            info["type"],
            ", ".join(sorted(info["categories"])),
            info["count"],
            info["example"],
            desc,
        ])

    auto_width(ws_fields, max_width=60)

    # ────────────────────────────────────────────────────────────────
    # SHEET 3: Platform Mapping Matrix
    # ────────────────────────────────────────────────────────────────
    ws_map = wb.create_sheet("Platform Mapping")
    headers = ["Source Field", "Data Type", "Target Platform Object",
                "Relationship Type", "Required/Optional", "Source Categories"]
    ws_map.append(headers)
    style_header(ws_map, 1, len(headers))

    for fp in sorted(field_catalog.keys()):
        leaf = fp.split(".")[-1].split("[")[0]
        if leaf in PLATFORM_OBJECTS:
            target, rel_type, req = PLATFORM_OBJECTS[leaf]
            info = field_catalog[fp]
            ws_map.append([
                fp,
                info["type"],
                target,
                rel_type,
                req,
                ", ".join(sorted(info["categories"])),
            ])

    auto_width(ws_map)

    # ────────────────────────────────────────────────────────────────
    # SHEET 4: ERD Relationships
    # ────────────────────────────────────────────────────────────────
    ws_erd = wb.create_sheet("Relationships (ERD)")
    headers = ["Source Entity", "Source Type", "Relationship",
                "Target Entity", "Target Type", "Cardinality", "Evidence"]
    ws_erd.append(headers)
    style_header(ws_erd, 1, len(headers))

    # Core ERD relationships derived from the data model
    erd_rows = [
        ("Firm", "Organization", "HAS_MANY", "Fund", "Vehicle", "1:N", "firm-id in fund features"),
        ("Firm", "Organization", "HAS_MANY", "Fund Family", "Vehicle Group", "1:N", "firm-id in family features"),
        ("Fund Family", "Vehicle Group", "CONTAINS", "Fund", "Vehicle", "1:N", "funds[] in fund-accounting"),
        ("Fund", "Vehicle", "HAS_MANY", "Investment", "Portfolio Company", "1:N", "fund_name in holdings rows"),
        ("Fund", "Vehicle", "HAS_MANY", "SPV", "Vehicle", "1:N", "entity_type=spv with fund features"),
        ("Fund", "Vehicle", "HAS_MANY", "Partner/LP", "Investor", "1:N", "partners in fund features"),
        ("Investment", "Portfolio Company", "HAS_MANY", "Security", "Holding", "1:N", "rows[] in holdings/shares"),
        ("Investment", "Portfolio Company", "HAS_MANY", "Valuation", "Assessment", "1:N", "post-money-list API"),
        ("Investment", "Portfolio Company", "HAS_MANY", "Option", "Security", "1:N", "rows[] in options API"),
        ("Security", "Holding", "OWNED_BY", "Fund/SPV", "Vehicle", "N:1", "owner_id, fund_name in row"),
        ("Security", "Holding", "HAS", "Vesting Schedule", "Schedule", "1:1", "has_vesting in holdings"),
        ("Security", "Holding", "HAS", "Transaction Receipt", "Document", "1:N", "transaction_receipts[]"),
        ("Investment", "Portfolio Company", "HAS", "Profile", "Metadata", "1:1", "corporation/profile API"),
        ("Investment", "Portfolio Company", "HAS", "Notes", "Document", "1:N", "notes/list API"),
        ("Investment", "Portfolio Company", "HAS", "Cap Table", "Financial", "1:1", "post-money-latest API"),
        ("Investment", "Portfolio Company", "HAS", "Tags", "Metadata", "1:N", "tags API"),
        ("Fund", "Vehicle", "HAS", "SOI Report", "Document", "1:1", "soi API"),
        ("Fund", "Vehicle", "HAS", "IRR Metrics", "Financial", "1:1", "deal-irr, explain-gross-irr"),
        ("Fund", "Vehicle", "HAS", "Capital Activity", "Financial", "1:N", "capital-activity API"),
        ("Fund", "Vehicle", "HAS", "Financial Statements", "Document", "1:N", "financials report API"),
        ("Fund", "Vehicle", "HAS", "Roll Forward", "Document", "1:1", "roll-forward report API"),
        ("Firm", "Organization", "HAS", "Entity List", "Catalog", "1:1", "gp-firm-entity-list API"),
        ("Firm", "Organization", "HAS", "Products", "Configuration", "1:1", "get-products API"),
        ("Investment", "Portfolio Company", "LINKED_TO", "Issuer Entity", "Legal Entity", "1:1", "issuer-entity-link in URLs"),
        ("Fund", "Vehicle", "HAS", "Fund Accounting", "Financial", "1:1", "fund-accounting features"),
    ]

    for r in erd_rows:
        ws_erd.append(list(r))

    auto_width(ws_erd)

    # ────────────────────────────────────────────────────────────────
    # SHEET 5: Extraction Summary
    # ────────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Extraction Summary")
    headers = ["Metric", "Value"]
    ws_sum.append(headers)
    style_header(ws_sum, 1, len(headers))

    # Aggregate manifest stats
    total_extracted = sum(m.get("summary", {}).get("total_extracted", 0) for m in manifests)
    total_failed = sum(m.get("summary", {}).get("total_failed", 0) for m in manifests)
    total_skipped = sum(m.get("summary", {}).get("total_skipped", 0) for m in manifests)

    summary_rows = [
        ("Total Export Runs", len(list(EXPORTS_ROOT.iterdir()))),
        ("Total JSON Files Analyzed", len(all_files)),
        ("Total Unique Entities", len(entities)),
        ("Total Unique Field Paths", len(field_catalog)),
        ("", ""),
        ("── Entity Breakdown ──", ""),
        ("Funds", len(classified.get("Fund", []))),
        ("SPVs", len(classified.get("SPV", []))),
        ("Portfolio Companies", len(classified.get("Portfolio Company", []))),
        ("Investors/LPs", len(classified.get("Investor/LP", []))),
        ("GP Entities", len(classified.get("GP Entity", []))),
        ("Other", len(classified.get("Other", []))),
        ("", ""),
        ("── Extraction Manifest ──", ""),
        ("Successful Extractions", total_extracted),
        ("Failed Extractions", total_failed),
        ("Skipped Endpoints", total_skipped),
        ("", ""),
        ("── Categories ──", ""),
    ]

    cat_counts = defaultdict(int)
    for f in all_files:
        cat_counts[f["category"]] += 1
    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        summary_rows.append((f"  {cat}", cnt))

    summary_rows.append(("", ""))
    summary_rows.append(("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    for r in summary_rows:
        ws_sum.append(list(r))
        if str(r[0]).startswith("──"):
            row_num = ws_sum.max_row
            ws_sum.cell(row=row_num, column=1).font = Font(bold=True, color="2F5496")

    auto_width(ws_sum)

    # ── Save ────────────────────────────────────────────────────────
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_FILE))
    print(f"\n[OK] Excel workbook saved to: {OUTPUT_FILE}")
    print(f"   Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
